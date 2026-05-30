from pathlib import Path
from collections import Counter, defaultdict
from openpyxl import load_workbook

DATA_ROOT = Path("data")

columns_counter = Counter()
columns_examples = defaultdict(list)

files = []
for ext in ("*.xlsx", "*.xlsm"):
    files.extend(DATA_ROOT.rglob(ext))

print(f"Found files: {len(files)}")

for file_path in files:
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"SKIP FILE {file_path}: {e}")
        continue

    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            rows = ws.iter_rows(values_only=True)
            headers = next(rows, None)

            if not headers:
                continue

            for h in headers:
                col = str(h or "").strip()

                if not col:
                    continue

                key = col.lower()
                columns_counter[key] += 1

                if len(columns_examples[key]) < 5:
                    columns_examples[key].append(
                        f"{file_path.parent.name} / {file_path.name} / {sheet_name}"
                    )

        except Exception as e:
            print(f"SKIP SHEET {file_path.name} / {sheet_name}: {e}")

print("\nTOP COLUMNS:\n")

for col, count in columns_counter.most_common(300):
    print(f"{count}x | {col}")
    for ex in columns_examples[col]:
        print(f"   - {ex}")