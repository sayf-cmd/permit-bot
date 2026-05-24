import re
from pathlib import Path
from difflib import SequenceMatcher
from openpyxl import load_workbook

DATA_ROOT = Path("data")

PHONE_RE = re.compile(r"(?:\+?\d[\d\s\-\(\)]{7,}\d)")

BUILDING_KEYS = ["building", "building name", "project", "property"]
UNIT_KEYS = ["unit", "unit number", "unit no", "plot", "plot pre reg no"]
DATE_KEYS = ["date", "regis", "transaction date", "procedure date"]
PRICE_KEYS = ["price", "value", "procedurevalue", "amount"]


def clean(value):
    return str(value or "").strip()


def norm(value):
    return clean(value).lower()


def similarity(a, b):
    return SequenceMatcher(None, norm(a), norm(b)).ratio()


def name_match(owner_name, row_text):
    owner = norm(owner_name)
    text = norm(row_text)

    if not owner or not text:
        return False

    if owner in text:
        return True

    owner_parts = [p for p in owner.split() if len(p) > 1]

    if len(owner_parts) >= 2:
        if all(part in text for part in owner_parts):
            return True

    for cell in text.split(" | "):
        if similarity(owner, cell) >= 0.82:
            return True

    return False


def find_by_keys(headers, row, keys):
    headers_norm = [norm(h) for h in headers]

    for key in keys:
        for i, h in enumerate(headers_norm):
            if key in h:
                if i < len(row):
                    value = clean(row[i])
                    if value:
                        return value

    return ""


def extract_phones(row):
    text = " | ".join(clean(v) for v in row)
    phones = []

    for match in PHONE_RE.findall(text):
        digits = re.sub(r"\D", "", match)

        if len(digits) < 7:
            continue

        if digits.startswith("971"):
            phone = "+" + digits
        elif digits.startswith("05") and len(digits) == 10:
            phone = "+971" + digits[1:]
        elif digits.startswith("5") and len(digits) == 9:
            phone = "+971" + digits
        else:
            phone = "+" + digits

        if phone not in phones:
            phones.append(phone)

    return phones


def row_to_dict(headers, row):
    result = {}

    for i, value in enumerate(row):
        key = clean(headers[i]) if i < len(headers) and headers[i] else f"col_{i + 1}"
        result[key] = clean(value)

    return result


def search_owner_everywhere(owner_name, max_results=50):
    results = []

    files = []
    for ext in ("*.xlsx", "*.xlsm"):
        files.extend(DATA_ROOT.rglob(ext))

    for file_path in files:
        source_folder = file_path.parent.name

        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
        except Exception as e:
            print(f"SKIP FILE {file_path}: {e}")
            continue

        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                rows = ws.iter_rows(values_only=True)

                headers = next(rows, None)
                if not headers:
                    continue

                headers = [clean(h) for h in headers]

                for excel_row_number, row in enumerate(rows, start=2):
                    row_values = [clean(v) for v in row]
                    row_text = " | ".join(row_values)

                    if not name_match(owner_name, row_text):
                        continue

                    result = {
                        "source_folder": source_folder,
                        "file_name": file_path.name,
                        "sheet_name": sheet_name,
                        "row_number": excel_row_number,
                        "building_name": find_by_keys(headers, row_values, BUILDING_KEYS),
                        "unit_number": find_by_keys(headers, row_values, UNIT_KEYS),
                        "owner_name": owner_name,
                        "phones": extract_phones(row_values),
                        "date": find_by_keys(headers, row_values, DATE_KEYS),
                        "price": find_by_keys(headers, row_values, PRICE_KEYS),
                        "full_row": row_to_dict(headers, row_values),
                    }

                    results.append(result)

                    if len(results) >= max_results:
                        return results

            except Exception as e:
                print(f"SKIP SHEET {file_path.name} / {sheet_name}: {e}")

    return results


def format_results_for_telegram(results):
    if not results:
        return "Ничего не найдено."

    messages = []

    for i, r in enumerate(results, start=1):
        phones = ", ".join(r["phones"]) if r["phones"] else "-"

        msg = (
            f"🔎 RESULT #{i}\n"
            f"Folder: {r['source_folder']}\n"
            f"File: {r['file_name']}\n"
            f"Sheet: {r['sheet_name']}\n"
            f"Row: {r['row_number']}\n\n"
            f"Owner: {r['owner_name']}\n"
            f"Building: {r['building_name'] or '-'}\n"
            f"Unit: {r['unit_number'] or '-'}\n"
            f"Date: {r['date'] or '-'}\n"
            f"Price: {r['price'] or '-'}\n"
            f"Phones: {phones}"
        )

        messages.append(msg)

    return "\n\n".join(messages)