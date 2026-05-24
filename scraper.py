import asyncio
import re
import time
import requests
from playwright.async_api import async_playwright
from openpyxl import load_workbook

# =========================
# CONFIG
# =========================

MAX_LISTINGS = 30

FILE_PATH = FILE_PATH = "/Users/wawirealestate/Desktop/Launcher.xlsx"
SHEET_NAME = "Leads"

BOT_TOKEN = "7986344001:AAEMjLIRPzDXzzJk5ryK0xiZv7ALPLO4pps"
CHAT_ID = "816494430"

semaphore = asyncio.Semaphore(3)


# =========================
# TELEGRAM
# =========================

def send_telegram(message):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    try:
        requests.post(url, data={
            "chat_id": CHAT_ID,
            "text": message
        })
    except:
        pass


# =========================
# FORMAT MESSAGE (CRM STYLE)
# =========================

def format_data_message(data):
    if not data:
        return "🏠 Property Overview\n\nNo new data found."

    msg = ""

    for item in data:
        msg += "🏠 Property Overview\n\n"

        msg += f"🏢 Permit Number: {item.get('permit','')}\n"
        msg += f"🏢 Building: {item.get('building','')}\n"
        msg += f"🏢 Unit Number: {item.get('unit_number','')}\n"
        msg += f"🛏️ Bedrooms: {item.get('beds','')}\n"
        msg += f"📐 Size: {item.get('size','')}\n"
        msg += f"💰 Price: {item.get('price','')}\n"
        msg += f"📅 Data Added: {item.get('date_added','')}\n\n"

        msg += "🔗 URL:\n"
        msg += f"{item.get('url','')}\n\n"

        msg += "⚡ Status: Completed\n"
        msg += "--------------------------\n\n"

    return msg


# =========================
# COLLECT URLS
# =========================

async def collect_urls(page):
    urls = set()

    for _ in range(8):
        links = page.locator("a[href*='/property/details-']")
        count = await links.count()

        for i in range(count):
            try:
                href = await links.nth(i).get_attribute("href")
                if href and "/property/details-" in href:
                    if href.startswith("http"):
                        urls.add(href)
                    else:
                        urls.add("https://www.bayut.com" + href)
            except:
                pass

        if len(urls) >= MAX_LISTINGS:
            break

        await page.mouse.wheel(0, 4000)
        await page.wait_for_timeout(700)

    return list(urls)[:MAX_LISTINGS]


# =========================
# PARSE WORKER (FAST)
# =========================

async def parse_worker(page, url):
    async with semaphore:
        try:
            await page.goto(url, wait_until="domcontentloaded")
            await page.wait_for_timeout(1200)

            body = await page.inner_text("body")

            price = ""
            beds = ""
            sqft = ""
            permit = ""

            m = re.search(r"AED[\s,]*([\d,]+)", body)
            if m:
                price = m.group(1)

            m = re.search(r"(\d+)\s*Beds?", body, re.IGNORECASE)
            if m:
                beds = m.group(1)
            elif "Studio" in body:
                beds = "Studio"

            m = re.search(r"([\d,]+)\s*(sqft|sq\.ft)", body, re.IGNORECASE)
            if m:
                sqft = m.group(1)

            m = re.search(r"(Permit Number|RERA Permit|Permit)\s*[:\-]?\s*(\d+)", body, re.IGNORECASE)
            if m:
                permit = m.group(2)

            if not permit:
                return None

            return {
                "permit": permit,
                "beds": beds,
                "price": price,
                "size": sqft,
                "url": url
            }

        except:
            return None


# =========================
# SCRAPER
# =========================

async def run_scraper():

    data = []

    building_name = input("Введи название билдинга: ").strip()

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            args=["--disable-blink-features=AutomationControlled"]
        )

        page = await browser.new_page()

        print("👉 Настрой фильтр на Bayut и нажми ENTER")
        input()

        urls = await collect_urls(page)
        print(f"Найдено: {len(urls)}")

        tasks = [parse_worker(page, url) for url in urls]
        results = await asyncio.gather(*tasks)

        data = [r for r in results if r]

        await browser.close()

    # =========================
    # EXCEL APPEND + ANTI DUPLICATE
    # =========================

    wb = load_workbook(FILE_PATH)
    ws = wb[SHEET_NAME]

    existing = set()

    for row in range(2, ws.max_row + 1):
        val = ws[f"B{row}"].value
        if val:
            existing.add(str(val))

    next_row = ws.max_row + 1

    new_data = []

    for item in data:
        permit = item["permit"]

        if permit in existing:
            continue

        ws[f"A{next_row}"] = ""  # date_added (if you use formula in Excel)
        ws[f"B{next_row}"] = permit
        ws[f"C{next_row}"] = building_name
        ws[f"E{next_row}"] = item["beds"]
        ws[f"F{next_row}"] = item["size"]
        ws[f"G{next_row}"] = item["price"]
        ws[f"I{next_row}"] = item["url"]

        existing.add(permit)
        next_row += 1

        new_data.append(item)

    wb.save(FILE_PATH)

    print(f"⚡ Добавлено: {len(new_data)} новых записей")

    # =========================
    # TELEGRAM REPORT
    # =========================

    message = format_data_message(new_data)
    send_telegram(message)

    print("📲 Telegram sent")


# =========================
# LOOP (EVERY 30 MIN)
# =========================

async def main_loop():
    while True:
        print("🚀 START CYCLE")

        try:
            await run_scraper()
        except Exception as e:
            print("ERROR:", e)

        print("😴 sleep 30 min")
        time.sleep(1800)


asyncio.run(main_loop())
