import json
import re
import sqlite3
from pathlib import Path

from openpyxl import load_workbook


DATA_ROOT = Path("data")
DB_PATH = Path("owners_index.db")

USEFUL_HEADER_WORDS = [
    "project", "building", "property", "master",
    "unit", "plot", "villa", "flat",
    "owner", "buyer", "seller", "name",
    "phone", "mobile", "contact",
]

PHONE_RE = re.compile(r"(?:\+?\d[\d\s\-\(\)]{7,}\d)")


def clean(value):
    return str(value or "").strip()


def norm(value):
    return clean(value).lower()


def row_has_phone(values):
    text = " | ".join(values)
    return bool(PHONE_RE.search(text))


def row_has_useful_data(headers, values):
    headers_text = " | ".join(norm(h) for h in headers)
    row_text = " | ".join(norm(v) for v in values)

    if row_has_phone(values):
        return True

    for word in USEFUL_HEADER_WORDS:
        if word in headers_text:
            # если колонка полезная и в строке есть данные
            if any(v.strip() for v in values):
                return True

    useful_words_in_row = [
        "damac", "emaar", "sobha", "nakheel",
        "azizi", "project", "building",
        "unit", "plot", "buyer", "seller",
        "owner"
    ]

    return any(word in row_text for word in useful_words_in_row)


def build_database():
    if DB_PATH.exists():
        DB_PATH.unlink()

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    cur.execute("""
        create table rows (
            id integer primary key autoincrement,
            source_folder text,
            file_name text,
            sheet_name text,
            row_number integer,
            row_text text,
            row_json text
        )
    """)

    cur.execute("create index idx_row_text on rows(row_text)")

    files = []
    for ext in ("*.xlsx", "*.xlsm"):
        files.extend(DATA_ROOT.rglob(ext))

    print(f"Found files: {len(files)}")

    inserted = 0
    skipped_empty = 0
    skipped_useless = 0

    for file_path in files:
        print(f"Reading: {file_path}")

        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
        except Exception as e:
            print(f"SKIP FILE {file_path}: {e}")
            continue

        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                rows = ws.iter_rows(values_only=True)

                headers = next(rows, None)
                if not headers:
                    continue

                headers = [
                    clean(h) or f"col_{i + 1}"
                    for i, h in enumerate(headers)
                ]

                empty_streak = 0

                for row_number, row in enumerate(rows, start=2):
                    values = [clean(v) for v in row]

                    if not any(values):
                        skipped_empty += 1
                        empty_streak += 1

                        if empty_streak >= 500:
                            break

                        continue

                    empty_streak = 0

                    if not row_has_useful_data(headers, values):
                        skipped_useless += 1
                        continue

                    row_dict = {}

                    for i, value in enumerate(values):
                        key = headers[i] if i < len(headers) else f"col_{i + 1}"
                        row_dict[key] = value

                    row_text = " | ".join(values).lower()

                    cur.execute(
                        """
                        insert into rows (
                            source_folder,
                            file_name,
                            sheet_name,
                            row_number,
                            row_text,
                            row_json
                        ) values (?, ?, ?, ?, ?, ?)
                        """,
                        (
                            file_path.parent.name,
                            file_path.name,
                            sheet_name,
                            row_number,
                            row_text,
                            json.dumps(row_dict, ensure_ascii=False),
                        )
                    )

                    inserted += 1

                    if inserted % 5000 == 0:
                        conn.commit()
                        print(
                            f"Inserted: {inserted} | "
                            f"Skipped empty: {skipped_empty} | "
                            f"Skipped useless: {skipped_useless}"
                        )

            except Exception as e:
                print(f"SKIP SHEET {file_path.name} / {sheet_name}: {e}")

    conn.commit()
    conn.close()

    print(f"Done. Inserted rows: {inserted}")
    print(f"Skipped empty: {skipped_empty}")
    print(f"Skipped useless: {skipped_useless}")
    print(f"Database created: {DB_PATH}")


if __name__ == "__main__":
    build_database()